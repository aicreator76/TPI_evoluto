from fastapi import Response
import io
import csv
import uuid
import time
import os
from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from .schemas import DpiRow, AncoraggioRow
from .models import Dpi, Ancoraggio
from .repository import upsert_many

MAX_BYTES = 5 * 1024 * 1024
router = APIRouter(prefix="/v1/cataloghi", tags=["cataloghi"])


def _jsonlog(level: str, **kv):  # log JSON semplice
    print({"level": level, **kv})


def _sanitize_csv_cell(v: str | None):
    if isinstance(v, str) and v and v[0] in "=+-@":
        return "'" + v
    return v


@router.get("/csv/template")
def csv_template():
    wb = Workbook()
    s1 = wb.active
    s1.title = "catalogo_dpi"
    s2 = wb.create_sheet("catalogo_ancoraggi")
    header = [
        "codice",
        "descrizione",
        "categoria",
        "prezzo_eur",
        "url",
        "_header_semver",
        "_note",
    ]
    for s in (s1, s2):
        s.append(header)
        s["F1"] = "1.0.0"
        s["G1"] = "Valori numerici con punto; URL http/https; codice [A-Z0-9-_]{3,32}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_cataloghi.xlsx"},
    )


@router.post("/csv/import")
async def csv_import(
    file: UploadFile = File(...),
    sheet: str | None = Query(
        default=None, description="Per XLSX: catalogo_dpi | catalogo_ancoraggi"
    ),
):
    t0 = time.time()
    cid = str(uuid.uuid4())
    raw = await file.read()
    if len(raw) > MAX_BYTES:
        raise HTTPException(status_code=413, detail="File troppo grande (>5MB)")

    rows, kind = [], None
    if file.filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[sheet or "catalogo_dpi"] if sheet else wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        data_iter = ([c.value for c in r] for r in ws.iter_rows(min_row=2))
        kind = "dpi" if ws.title == "catalogo_dpi" else "ancoraggi"
        for r in data_iter:
            item = dict(zip(header, r))
            rows.append(item)
    else:
        text = raw.decode("utf-8-sig", errors="ignore").splitlines()
        if not text:
            raise HTTPException(status_code=400, detail="File vuoto")
        header = next(csv.reader([text[0]]))
        kind = "dpi" if "dpi" in (sheet or "dpi").lower() else "ancoraggi"
        for r in csv.DictReader(text):
            rows.append(r)

    expect = {"codice", "descrizione", "categoria", "prezzo_eur", "url"}
    if not expect.issubset({(h or "").lower() for h in header}):
        raise HTTPException(status_code=400, detail=f"Header mancante: richiesti {sorted(expect)}")

    valid, invalid = [], []

    def coerce(d: dict) -> dict:
        m = {(k or "").lower(): d.get(k) for k in header}
        m["prezzo_eur"] = (
            float(m["prezzo_eur"]) if (m.get("prezzo_eur") not in (None, "")) else None
        )
        if m.get("url"):
            m["url"] = str(m["url"])
        m["codice"] = _sanitize_csv_cell(str(m["codice"])) if m.get("codice") else m["codice"]
        return m

    for r in rows:
        try:
            m = coerce(r)
            (DpiRow if kind == "dpi" else AncoraggioRow)(**m)
            valid.append(m)
        except Exception as e:
            invalid.append((r, str(e)))

    engine = create_engine(os.getenv("DATABASE_URL", "sqlite:///./app.db"), future=True)
    Session = sessionmaker(engine, expire_on_commit=False, future=True)
    os.makedirs("./artifacts/rejects", exist_ok=True)
    reject_path = f"./artifacts/rejects/reject_{cid}.csv"

    inserted = updated = skipped = 0
    with Session.begin() as s:
        if kind == "dpi":
            i, u, k = upsert_many(s, Dpi, valid)
        else:
            i, u, k = upsert_many(s, Ancoraggio, valid)
        inserted += i
        updated += u
        skipped += k

    if invalid:
        with open(reject_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["motivo", "riga"])
            for r, why in invalid:
                w.writerow([why, r])

    dur_ms = int((time.time() - t0) * 1000)
    _jsonlog(
        "INFO",
        event="cataloghi_import",
        correlationId=cid,
        metrics={"duration_ms": dur_ms, "invalid": len(invalid)},
        kind=kind,
    )

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "invalid": len(invalid),
        "rejectLog": (reject_path if invalid else None),
        "correlationId": cid,
        "metrics": {"duration_ms": dur_ms, "batch": len(valid)},
    }


@router.head("/csv/template")
def csv_template_head():
    return Response(
        status_code=200,
        headers={
            "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        },
    )
